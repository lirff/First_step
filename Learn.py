from openpyxl import Workbook
wb = Workbook()
ws = wb.active
WorkSheet1 = wb.create_sheet("WorkSheet1") # вставляет его в конец по умолчанию
WorkSheet2 = wb.create_sheet("WorkSheet2", 0) # ставит на первое место
print(wb.sheetnames)
nomecell = str(input('Введите записываемую ячейку: '))
cell = WorkSheet1[nomecell]
WorkSheet1[nomecell] = str(input('Введите данные в а4: '))
wb.save("Testing_book.xlsx")